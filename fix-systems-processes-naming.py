'''
This script is to fix the names of the systems processes, so that when i upload 
them to OpenLCA they don't overwrite the unit processes.

I want to:
    1. from `scraper-output` open each .xlsx file with the suffix " S.xlsx"
    2. in each file, open the first sheet, titled "General information"
    3. find the cell B3 (cell A3 is "Name")
    4. if the text in cell B3 ends with " U" replace with " S", otherwise, append " S"
'''

import openpyxl
import os

def stackoverflow(files):
    # From this stackoverflow post https://stackoverflow.com/questions/58075018/change-contents-of-a-specific-cell-in-multiple-xlsx-files-in-python
    count = 0
    for f in files:
        if f[-4:] == "xlsx":
            book = openpyxl.load_workbook(f)
            sheet = book.active
            for row in sheet.iter_rows(values_only=True):
                for data in row:
                    if data == "value1.n":
                        count = count + 1
                        print(data)
                        print(count)

def find_cell(files):
    # From this stackoverflow post https://stackoverflow.com/questions/58075018/change-contents-of-a-specific-cell-in-multiple-xlsx-files-in-python
    count = 0
    for f in files:
        # 1. from `scraper-output` open each .xlsx file with the suffix " S.xlsx"
        if f[-7:] == " S.xlsx":
            print("Fixing file:",f)
            try:
                book = openpyxl.load_workbook(f)
                # 2. in each file, open the first sheet, titled "General information"
                # sheet = book.get_sheet_by_name("General information") # depreciated function
                sheet = book["General information"]
                # 3. find the cell B3 (cell A3 is "Name")
                name = sheet.cell(3,2)
                print("Name:",name.value)
                value = name.value
                # 4. if the text in cell B3 ends with " U" replace with " S", otherwise, append " S"
                if value[-2:] == " S":
                    print("\nALREADY GOOD\n")
                else:
                    if value[-2:] == " U":
                        value = value[:-2] + " S"
                    elif value[-2:] != " S":
                        value = value + " S"
                    print(value)
                    # for row in sheet.iter_rows(values_only=True):
                    #     for data in row:
                    #         # if data == "value1.n":
                    #         if data == "Name":
                    #             count = count + 1
                    #             print(data)
                    #             print(count)
                    sheet.cell(3,2,value=value)
                    book.save(f)
                    print("saved file! : ",f, "\n")
            except:
                print("error for file", f)




if __name__ == '__main__':
    print(os.getcwd())
    
    ## single folder
    # folder = 'scraper-output/transport/water/' # It worked on the water ones!!
    # folder = 'scraper-output/materials/textiles' # worked!
    # wood - done
    # water - done
    # plastics - done
    # minerals - done
    # fuels - done
    # construction - done
    # chemicals - done
    # agriculture, in progress through the mega business, and also started a smaller one, to see what is faster..... fingers crossed we don't corrupt anything lol
    folder = 'scraper-output/materials/agriculture'
    files = os.listdir(folder)
    
    ## Multi folder
    # path = 'scraper-output/waste'
    # path = 'scraper-output/transport'
    # path = 'scraper-output/energy'
    # path = 'scraper-output/processing'
    # path = 'scraper-output/materials'
    # folders = os.listdir(path=path)
    # files = []
    # for f in folders:
    #     fs = os.listdir(os.path.join(path,f))
    #     fs = [os.path.join(f,file) for file in fs]
    #     files = files + (fs)
    # files = [os.listdir(os.path.join(path,f)) for f in folders]

    print(files)
    os.chdir(folder)
    print(os.getcwd())


    find_cell(files=files)

'''
Just checking on the progress of the scraper, at 4:51pm 
the materials agriculture one is on 
"downloading ('culled breeders, Annual Rainfall more than 450mm, WA, at farm AU S', 'https://www.auslci.com.au/Datasets/ExcelFiles/culled%20breeders,%20Annual%20Rainfall%20more%20than%20450mm,%20WA,%20at%20farm%20AU%20S.xlsx')"

and the processing agriculture one is on
downloading ('ripping, medium implement, horticulture AU S', 'https://www.auslci.com.au/Datasets/ExcelFiles/ripping,%20medium%20implement,%20horticulture%20AU%20S.xlsx') 
 to  scraper-output/processing/agricultural/

I think they've been there for a while..... i might move them on, and download manually. In 5 mins


5:03, i cut off processing agriculture, it was on
downloading ('rolling, cotton AU S', 'https://www.auslci.com.au/Datasets/ExcelFiles/rolling,%20cotton%20AU%20S.xlsx') 
 to  scraper-output/processing/agricultural/

5:04 i cut off materials agriculture
downloading ('culled breeders, Australia AU S', 'https://www.auslci.com.au/Datasets/ExcelFiles/culled%20breeders,%20Australia%20AU%20S.xlsx')

'''